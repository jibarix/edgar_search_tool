"""Inventory every sheet in the CapIQ comparables workbook: list each
sheet's column headers (the per-period metrics) so we can decide which
ones to replicate in the metrics engine, defer, or ignore.

For each sheet:
  - finds the header row (first row whose first cell is "Company Name",
    "Ticker", or starts with a known metric label)
  - dumps the column labels under that header
  - flags whether the sheet appears to be a "wide" per-company sheet
    (rows are companies) or a "narrow" detail sheet
"""
from __future__ import annotations

import sys, io, os
os.environ.setdefault("PYTHONIOENCODING", "utf-8")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from pathlib import Path
from openpyxl import load_workbook

SRC = Path(
    r"C:\Users\arroy\OneDrive - Cabrera Auto\valuations\data\comparables"
    r"\comparables_dealerships_20251231.xlsx"
)

HEADER_HINTS = ("Company Name", "Ticker", "Company")


def find_header_row(rows: list[tuple]) -> int | None:
    for i, r in enumerate(rows):
        if not r:
            continue
        first = str(r[0]).strip() if r[0] else ""
        if first in HEADER_HINTS:
            return i
    return None


def main():
    wb = load_workbook(SRC, data_only=True, read_only=True)
    print(f"File: {SRC.name}")
    print(f"Sheets: {len(wb.sheetnames)}\n")

    for sname in wb.sheetnames:
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        non_empty = sum(1 for r in rows if r and any(c is not None for c in r))
        h = find_header_row(rows)
        print(f"=== {sname}  ({ws.max_row} rows, {ws.max_column} cols, "
              f"{non_empty} non-empty) ===")

        if h is None:
            # Likely a narrative / footer / freeform sheet
            print("  (no Company Name / Ticker header — narrative sheet?)")
            # Show the first few non-empty cells anyway
            shown = 0
            for r in rows:
                if not r or not any(c is not None for c in r):
                    continue
                cells = [str(c).strip() for c in r if c is not None]
                if cells:
                    print(f"  > {cells[0][:80]}")
                    shown += 1
                if shown >= 5:
                    break
            print()
            continue

        hdr = [str(c).strip() if c else "" for c in rows[h]]
        # Trim trailing empty columns
        while hdr and hdr[-1] == "":
            hdr.pop()
        # Count companies in body
        body_rows = sum(
            1 for r in rows[h + 1:] if r and r[0] and str(r[0]).strip()
        )
        print(f"  header at row {h}, {body_rows} body rows, "
              f"{len(hdr)} cols")
        print(f"  columns:")
        for i, label in enumerate(hdr):
            if not label:
                continue
            print(f"    [{i:2d}] {label}")
        print()


if __name__ == "__main__":
    main()
