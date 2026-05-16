"""Inspect the Capital IQ comparables_dealerships_20251231.xlsx (re-saved as xlsx)."""
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\arroy\OneDrive - Cabrera Auto\valuations\data\comparables\comparables_dealerships_20251231.xlsx")
print(f"File: {SRC}  exists: {SRC.exists()}  size: {SRC.stat().st_size}")

wb = load_workbook(SRC, data_only=True, read_only=True)
print(f"Sheets: {wb.sheetnames}")

for sname in wb.sheetnames:
    ws = wb[sname]
    print(f"\n=== Sheet: {sname}  dims: {ws.max_row} rows x {ws.max_column} cols ===")
    for ri, row in enumerate(ws.iter_rows(values_only=True)):
        if ri >= 50:
            print("  ... (truncated to 50 rows)")
            break
        cells = [str(c).strip() if c is not None else "" for c in row]
        # Trim trailing empties
        while cells and cells[-1] == "":
            cells.pop()
        if not cells:
            continue
        preview = [c[:50] for c in cells]
        print(f"  row {ri}: {preview}")
