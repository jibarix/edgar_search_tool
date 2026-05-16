"""Shared comp-set profiles for the Capital IQ reconcile harness.

The reconcile_capiq_* scripts originally hardcoded one comp set (the 8
US auto dealers). A "profile" parameterises everything that varies
between comp files so the same scripts run against any Capital IQ
workbook:

    capiq_path        the workbook (canonical OneDrive location)
    as_of             the LTM / pricing anchor date
    tickers           [(ticker, display name), ...] — only US SEC filers
                      that resolve in EDGAR; foreign listings are skipped
    extension_rules   industry company-extension XBRL rules
                      (DEALER_RULES for dealers; [] for REITs)
    chain_overrides   per-comp-set concept fallback-chain replacements,
                      threaded into NormalizedStatement. Adapts concept
                      resolution to industry tagging without a global
                      reorder that would regress other comp sets.

Select with `--profile <name>` (default `dealers`, preserving the
original behaviour). See docs for the REIT investigation that motivated
`chain_overrides`.
"""
from __future__ import annotations

import argparse
import os
import re
import subprocess
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from edgar._extension_mappings import DEALER_RULES
from utils.cache import Cache

# Every reconcile script passes through get_profile() exactly once at
# startup, so it is the single chokepoint for the deferred-eviction
# maintenance sweep (KI-5 step 3): the cache read path no longer
# unlinks expired entries, so something must reclaim them. These are
# the namespaces the engine writes during a harness run.
_HARNESS_CACHE_NAMESPACES = (
    "company_lookup",
    "filing_data",
    "xbrl_instance",
    "xbrl_parser",
)
_caches_swept = False


def sweep_caches() -> None:
    """Reclaim expired/corrupt entries across the harness cache dirs.

    Idempotent and best-effort: runs once per process, and a locked or
    missing namespace is skipped rather than raised (cleanup() is
    itself race-tolerant — see KI-5). expiry at construction is
    irrelevant here; cleanup() reads each entry's own expires_at.
    """
    global _caches_swept
    if _caches_swept:
        return
    _caches_swept = True
    for ns in _HARNESS_CACHE_NAMESPACES:
        try:
            Cache(ns).cleanup()
        except Exception:
            pass

# `Name (Exchange:Ticker)` — column-A convention on every CapIQ sheet.
TICKER_RE = re.compile(r"\(([^():]+):([^()]+)\)\s*$")

_COMPARABLES_DIR = Path(
    r"C:\Users\arroy\OneDrive - Cabrera Auto\valuations\data\comparables"
)


def to_float(v):
    """CapIQ cell -> float, mapping its sentinels ('-', 'NM', 'NA') to None."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "NM", "NA"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


@dataclass(frozen=True)
class CompProfile:
    name: str
    capiq_path: Path
    as_of: str
    tickers: list[tuple[str, str]]
    extension_rules: list = field(default_factory=list)
    chain_overrides: dict = field(default_factory=dict)


# ── Dealers (original comp set) ───────────────────────────────────────

DEALERS = CompProfile(
    name="dealers",
    capiq_path=_COMPARABLES_DIR / "comparables_dealerships_20251231.xlsx",
    as_of="2025-12-31",
    tickers=[
        ("ABG", "Asbury Automotive Group"),
        ("AN", "AutoNation"),
        ("CRMT", "America's Car-Mart"),
        ("LAD", "Lithia Motors"),
        ("KMX", "CarMax"),
        ("PAG", "Penske Automotive Group"),
        ("GPI", "Group 1 Automotive"),
        ("SAH", "Sonic Automotive"),
    ],
    extension_rules=DEALER_RULES,
    chain_overrides={},
)


# ── Shopping-mall / real-estate comp set ──────────────────────────────
# Only 4 of the 32 names are US SEC filers that resolve in EDGAR; the
# other 28 are foreign listings. KW tags neither Revenues nor
# OperatingIncomeLoss (a real-estate *investment* company — CapIQ's
# "revenue" reconstructs investment income that isn't GAAP operating
# revenue) and ABCP is a shell; both are kept for completeness but are
# expected to be structurally non-reconcilable.
#
# REIT/LP filers tag rental income under Revenues / OperatingLeaseLeaseIncome
# and additionally carry a tiny RevenueFromContractWithCustomer line. The
# global chain (contract-tag first, correct for dealers) would pick that
# tiny line, so this override flips the priority for this comp set only.
MALLS = CompProfile(
    name="malls",
    capiq_path=(
        _COMPARABLES_DIR
        / "Company Comparable Analysis Shopping Malls_20260407.xlsx"
    ),
    as_of="2026-04-07",
    tickers=[
        ("KW", "Kennedy-Wilson Holdings"),
        ("ARL", "American Realty Investors"),
        ("NEN", "New England Realty Associates LP"),
        ("ABCP", "AmBase Corporation"),
    ],
    extension_rules=[],
    chain_overrides={
        "revenue": [
            ("Revenue", "Revenues"),
            ("Revenue", "RevenueFromContractWithCustomerExcludingAssessedTax"),
            ("Revenue", "RevenueFromContractWithCustomerIncludingAssessedTax"),
            ("Income", "OperatingLeaseLeaseIncome"),
            ("Revenue", "SalesRevenueNet"),
        ],
    },
)


# ── Health Care Services screen (Company Screening Report) ────────────
# NOT the 11-sheet comparables template the other profiles consume. This
# is a CapIQ *Company Screening Report*: sheets ['Screening',
# 'Aggregates', 'Screen Criteria'], one row per company on 'Screening'
# with a dedicated `Exchange:Ticker` column (no `Name (Exch:Tkr)` key)
# and raw $USDmm LTM fundamentals + 5Y β/R². Only the screening-aware
# script (reconcile_capiq_screening.py) reads it; the comparables
# reconcile scripts iterate `tickers`, which is empty here, so running
# them against this profile is a no-op rather than a crash.
#
# The file ships no extraction date, so `as_of` is the most recent
# completed quarter-end before today (CapIQ "[Latest]"/"[LTM]" anchor);
# override with --as-of if a different snapshot is needed.
_REPO_ROOT = Path(__file__).resolve().parent.parent

SCREENING_HC = CompProfile(
    name="screening_hc",
    capiq_path=_REPO_ROOT / "data" / "bottomUpBeta.xlsx",
    as_of="2026-03-31",
    tickers=[],
    extension_rules=[],
    # Health-care multi-segment filers (e.g. CHE = VITAS hospice +
    # Roto-Rooter) tag a partial ASC606 contract-revenue line AND a
    # total `Revenues` line; the global chain is contract-tag-first
    # (correct for dealers) and would pick the partial — CHE FY2024
    # 1,531M vs the 2,431M `Revenues` total. Flip to `Revenues`-first
    # for this screen only, same rationale/shape as the MALLS override.
    # `Revenues` is the broadest GAAP total-revenue line; chain falls
    # through to the contract tags for names that only tag those.
    chain_overrides={
        "revenue": [
            ("Revenue", "Revenues"),
            ("Revenue", "RevenueFromContractWithCustomerExcludingAssessedTax"),
            ("Revenue", "RevenueFromContractWithCustomerIncludingAssessedTax"),
            ("Revenue", "SalesRevenueNet"),
            ("Revenue", "SalesRevenueGoodsNet"),
        ],
    },
)


PROFILES: dict[str, CompProfile] = {
    p.name: p for p in (DEALERS, MALLS, SCREENING_HC)
}


def get_profile(name: str) -> CompProfile:
    # Harness startup chokepoint: reclaim expired cache entries before
    # the run so the deferred-eviction design (KI-5) actually frees
    # disk. Runs once per process; never fails the run.
    sweep_caches()
    try:
        return PROFILES[name]
    except KeyError:
        raise SystemExit(
            f"unknown profile {name!r}; choose one of {sorted(PROFILES)}"
        )


def select_profile(description: str) -> CompProfile:
    """Parse `--profile NAME` from argv. Defaults to `dealers`."""
    ap = argparse.ArgumentParser(description=description)
    ap.add_argument(
        "--profile", default="dealers", choices=sorted(PROFILES),
        help="comp-set profile (default: dealers)",
    )
    args, _ = ap.parse_known_args()
    return get_profile(args.profile)


def load_capiq_workbook(path: Path):
    """`load_workbook` with an OneDrive Files-On-Demand fallback.

    The dealer file is hydrated locally and opens directly. Newly
    "dropped" files (e.g. the mall workbook) are cloud-only placeholders
    that raise PermissionError/OSError when openpyxl mmaps them. Force
    hydration with a full-byte PowerShell read, copy to %TEMP%, retry.
    """
    try:
        return load_workbook(path, data_only=True, read_only=True)
    except (PermissionError, OSError):
        local = Path(tempfile.gettempdir()) / f"capiq_{path.name}"
        ps = (
            f"$null = Get-Content -LiteralPath '{path}' -Encoding Byte "
            f"-ReadCount 0; "
            f"Copy-Item -LiteralPath '{path}' -Destination '{local}' -Force"
        )
        subprocess.run(
            ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
            check=True, capture_output=True, text=True,
        )
        return load_workbook(local, data_only=True, read_only=True)
