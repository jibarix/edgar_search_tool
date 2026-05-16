"""Extract the full company list from the Capital IQ dealership comps file.

Output: parse company names + tickers, split US-listed (EDGAR-reachable) from foreign.
"""
import json
import re
from pathlib import Path
from openpyxl import load_workbook

SRC = Path(r"C:\Users\arroy\OneDrive - Cabrera Auto\valuations\data\comparables\comparables_dealerships_20251231.xlsx")
OUT = Path(__file__).resolve().parent.parent / "data" / "capiq_dealership_comps.json"

wb = load_workbook(SRC, data_only=True, read_only=True)
ws = wb["Financial Data"]

# Row 13 has the header, rows 14+ have companies. Capital IQ format:
# "<Company Name>, Inc. (NYSE:ABG)"   or   "<Foreign Co> (TSE:XXXX)"
# Match the LAST (exch:ticker) bracket — names can contain other parens
# (e.g. "Bilia AB (publ) (OM:BILI A)").
NAME_RE = re.compile(r"^(?P<name>.+)\s*\((?P<exch>[^():]+):(?P<ticker>[^()]+)\)\s*$")

rows = list(ws.iter_rows(values_only=True))
print(f"Total rows: {len(rows)}")

companies = []
for r in rows[13:]:  # skip header
    cell = r[0]
    if not cell:
        continue
    s = str(cell).strip()
    if not s:
        continue
    m = NAME_RE.match(s)
    if m:
        companies.append({
            "raw": s,
            "name": m.group("name").strip(),
            "exchange": m.group("exch").strip(),
            "ticker": m.group("ticker").strip(),
        })
    else:
        # No ticker bracketed — probably private/delisted
        companies.append({"raw": s, "name": s, "exchange": None, "ticker": None})

print(f"Parsed companies: {len(companies)}")

# Split by US-listed exchanges
US_EXCHANGES = {"NYSE", "NasdaqGS", "NasdaqGM", "NasdaqCM", "NASDAQ", "AMEX", "OTCMKTS"}
us = [c for c in companies if c["exchange"] in US_EXCHANGES]
foreign = [c for c in companies if c["exchange"] and c["exchange"] not in US_EXCHANGES]
unknown = [c for c in companies if not c["exchange"]]

print(f"\nUS-listed (EDGAR-reachable): {len(us)}")
for c in us:
    print(f"  {c['exchange']}:{c['ticker']:<6} {c['name']}")

print(f"\nForeign-listed: {len(foreign)}")
exch_counts: dict[str, int] = {}
for c in foreign:
    exch_counts[c['exchange']] = exch_counts.get(c['exchange'], 0) + 1
for ex, n in sorted(exch_counts.items(), key=lambda x: -x[1]):
    print(f"  {ex:<10} x{n}")

print(f"\nUnparsed / private: {len(unknown)}")
for c in unknown:
    print(f"  {c['raw'][:80]}")

OUT.write_text(json.dumps({
    "as_of": "2025-12-31",
    "source": str(SRC),
    "total": len(companies),
    "us": us,
    "foreign": foreign,
    "unknown": unknown,
}, indent=2))
print(f"\nWrote: {OUT}")
